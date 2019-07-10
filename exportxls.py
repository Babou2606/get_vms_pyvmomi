from openpyxl import Workbook

def exportxls(children):
    print("Export des informations sous format xlsx")
    path = r"fichier.xlsx"

    i = 2 #Commencer à la seconde ligne
    classeur = Workbook()# On ajoute une feuille au classeur
    feuille = classeur.active
    feuille.title = "FCC"

    for virtual_machine in children:
        summary = virtual_machine.summary

        a = feuille.cell(row=1, column=1)
        a.value = "Nom"
        b = feuille.cell(row=1, column=2)
        b.value = "Adresse IP"
        c = feuille.cell(row=1, column=3)
        c.value = "Système d'exploitation"
        d = feuille.cell(row=1, column=4)
        d.value = "Etat"
        e = feuille.cell(row=1, column=5)
        e.value = "VMware-tools"
        f = feuille.cell(row=1, column=6)
        f.value = "Chemin de la VM"

        if summary.config.template == False:
            a=feuille.cell(row=i, column=1)
            a.value=summary.config.name
            b=feuille.cell(row=i, column=2)
            b.value=summary.guest.ipAddress
            c=feuille.cell(row=i, column=3)
            c.value=summary.guest.guestFullName
            d=feuille.cell(row=i, column=4)
            d.value=str(summary.runtime.powerState)


            tools_version = summary.guest.toolsStatus
            if tools_version is not None:
                e = feuille.cell(row=i, column=5)
                e.value = str(tools_version)
            f = feuille.cell(row=i, column=6)
            f.value=summary.config.vmPathName
            i = i + 1


    column_count = feuille.max_column
    feuille.column_dimensions['A'].width = 55
    feuille.column_dimensions['B'].width = 25
    feuille.column_dimensions['C'].width = 55
    feuille.column_dimensions['D'].width = 15
    feuille.column_dimensions['E'].width = 15
    feuille.column_dimensions['F'].width = 55

    print(column_count)

    classeur.save(path)
    print("Fichier créé")